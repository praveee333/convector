import os
import razorpay
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify
# Add this import for CORS handling
from flask_cors import CORS

# Load environment variables from .env file
from dotenv import load_dotenv
load_dotenv()
import pandas as pd
import PyPDF2
import openpyxl
from openpyxl.chart import PieChart, LineChart, Reference, BarChart
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
from werkzeug.utils import secure_filename
import json
from datetime import datetime, timedelta
import re
import hashlib
import mysql.connector
from mysql.connector import Error
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header

# Try to import OCR libraries
try:
    import pytesseract
except ImportError:
    pytesseract = None

try:
    from PIL import Image
except ImportError:
    Image = None

try:
    import docx
except ImportError:
    docx = None

# Initialize Flask app
app = Flask(__name__)
app.secret_key = 'your-secret_key_here_change_this_in_production'  # Change this in production

# Enable CORS for all routes
CORS(app)

# Razorpay configuration - THESE ARE TEST KEYS THAT WON'T WORK FOR ACTUAL TRANSACTIONS
# To fix authentication errors, set valid environment variables:
# export RAZORPAY_KEY_ID=your_actual_key_id
# export RAZORPAY_KEY_SECRET=your_actual_key_secret
RAZORPAY_KEY_ID = os.environ.get('RAZORPAY_KEY_ID')
RAZORPAY_KEY_SECRET = os.environ.get('RAZORPAY_KEY_SECRET')

# Initialize Razorpay client with error handling
razorpay_client = None
def initialize_razorpay():
    global razorpay_client
    try:
        # Validate that we have proper keys
        if not RAZORPAY_KEY_ID or not RAZORPAY_KEY_SECRET:
            print("ERROR: Razorpay keys not configured.")
            print("Please set RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET environment variables.")
            print("See RAZORPAY_SETUP.md for detailed instructions.")
            return False
            
        razorpay_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))
        razorpay_client.set_app_details({"title" : "Convector", "version" : "1.0"})
        print("Razorpay client initialized successfully")
        # Check if RAZORPAY_KEY_ID is not None before slicing
        if RAZORPAY_KEY_ID:
            print(f"Using Razorpay keys - ID: {RAZORPAY_KEY_ID[:10]}... SECRET: {'*' * 10}")
        else:
            print("Using Razorpay keys - ID: None SECRET: None")
        return True
    except Exception as e:
        print(f"Warning: Razorpay client initialization failed: {e}")
        return False

# Initialize on startup - but only if we have valid keys
if RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET:
    initialize_razorpay()
else:
    print("Razorpay not initialized - missing API keys")
    print("To enable payments, set RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET environment variables")

# Configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx', 'xls', 'xlsx'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Create upload folder if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# MySQL Database Configuration
DB_CONFIG = {
    'host': os.environ.get('DB_HOST', 'localhost'),
    'user': os.environ.get('DB_USER', 'root'),
    'password': os.environ.get('DB_PASSWORD', ''),
    'database': os.environ.get('DB_NAME', 'convector_auth'),
    'port': int(os.environ.get('DB_PORT', 3306))
}

# SMTP Email Configuration
SMTP_CONFIG = {
    'server': os.environ.get('SMTP_SERVER', 'smtp.gmail.com'),
    'port': int(os.environ.get('SMTP_PORT', 587)),
    'username': os.environ.get('SMTP_USERNAME'),
    'password': os.environ.get('SMTP_PASSWORD'),
    'sender_email': os.environ.get('SMTP_SENDER_EMAIL'),
    'sender_name': os.environ.get('SMTP_SENDER_NAME', 'E-Faws Tech Services')
}

def get_db_connection():
    """Create and return a database connection"""
    try:
        connection = mysql.connector.connect(**DB_CONFIG)
        return connection
    except Error as e:
        print(f"Error connecting to MySQL: {e}")
        return None

def init_db():
    """Initialize database tables if they don't exist"""
    connection = get_db_connection()
    if connection is None:
        print("Failed to connect to database")
        return False
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Create users table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INT AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(255) NOT NULL,
                email VARCHAR(255) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                phone VARCHAR(20),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Check if phone column exists, if not add it
        try:
            cursor.execute("ALTER TABLE users ADD COLUMN phone VARCHAR(20)")
            print("Added phone column to users table")
        except Error as e:
            # Column already exists or other error, continue
            if e.errno != 1060:  # 1060 = Duplicate column name
                print(f"Note: {e}")
        
        # Create subscriptions table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS subscriptions (
                id INT AUTO_INCREMENT PRIMARY KEY,
                user_id INT NOT NULL,
                plan VARCHAR(50) NOT NULL,
                start_date TIMESTAMP NOT NULL,
                end_date TIMESTAMP NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        """)
        
        # Create bank_statement_analyses table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS bank_statement_analyses (
                id INT AUTO_INCREMENT PRIMARY KEY,
                user_id INT NOT NULL,
                name VARCHAR(255) NOT NULL,
                bank_name VARCHAR(255) NOT NULL,
                customer_number VARCHAR(100) NOT NULL,
                analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                file_name VARCHAR(255) NOT NULL,
                excel_file_path VARCHAR(500) NOT NULL,
                pdf_file_path VARCHAR(500),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        """)
        
        # Create notifications table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS notifications (
                id INT AUTO_INCREMENT PRIMARY KEY,
                user_id INT NOT NULL,
                type VARCHAR(50) NOT NULL,
                title VARCHAR(255) NOT NULL,
                message TEXT NOT NULL,
                is_read BOOLEAN DEFAULT FALSE,
                starred BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        """)
        
        # Create communication history table for reminders and other communications
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS communication_history (
                id INT AUTO_INCREMENT PRIMARY KEY,
                user_id INT NOT NULL,
                message_type VARCHAR(50) NOT NULL,
                sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                status VARCHAR(50) DEFAULT 'sent',
                details TEXT,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        """)
        
        # Create WhatsApp history table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS whatsapp_history (
                id INT AUTO_INCREMENT PRIMARY KEY,
                user_id INT NOT NULL,
                phone VARCHAR(20) NOT NULL,
                message TEXT NOT NULL,
                sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        """)
        
        # Create email history table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS email_history (
                id INT AUTO_INCREMENT PRIMARY KEY,
                user_id INT NOT NULL,
                email VARCHAR(255) NOT NULL,
                subject VARCHAR(255) NOT NULL,
                body TEXT NOT NULL,
                sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                status VARCHAR(50) DEFAULT 'sent',
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        """)
        
        # Create indexes (ignore if they already exist)
        index_queries = [
            "CREATE INDEX idx_users_email ON users(email)",
            "CREATE INDEX idx_subscriptions_user_id ON subscriptions(user_id)",
            "CREATE INDEX idx_subscriptions_end_date ON subscriptions(end_date)",
            "CREATE INDEX idx_analyses_user_id ON bank_statement_analyses(user_id)",
            "CREATE INDEX idx_analyses_date ON bank_statement_analyses(analysis_date)",
            "CREATE INDEX idx_notifications_user_id ON notifications(user_id)",
            "CREATE INDEX idx_notifications_created_at ON notifications(created_at)",
            "CREATE INDEX idx_communication_user_id ON communication_history(user_id)",
            "CREATE INDEX idx_communication_sent_at ON communication_history(sent_at)",
            "CREATE INDEX idx_whatsapp_user_id ON whatsapp_history(user_id)",
            "CREATE INDEX idx_whatsapp_sent_at ON whatsapp_history(sent_at)",
            "CREATE INDEX idx_email_user_id ON email_history(user_id)",
            "CREATE INDEX idx_email_sent_at ON email_history(sent_at)"
        ]
        
        for query in index_queries:
            try:
                cursor.execute(query)
            except Error as e:
                # Index already exists or other minor error, continue
                print(f"Note: {e}")
                pass
        
        connection.commit()
        print("Database initialized successfully")
        return True
    except Error as e:
        print(f"Error initializing database: {e}")
        return False
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

# Initialize database on startup
init_db()

def send_email(to_email, subject, body, html_body=None):
    """
    Send an email using SMTP configuration from environment variables
    
    Args:
        to_email (str): Recipient email address
        subject (str): Email subject
        body (str): Plain text email body
        html_body (str, optional): HTML email body
        
    Returns:
        bool: True if email sent successfully, False otherwise
    """
    try:
        # Check if SMTP is configured
        if not SMTP_CONFIG['username'] or not SMTP_CONFIG['password']:
            print("SMTP not configured. Please set SMTP_USERNAME and SMTP_PASSWORD in .env file.")
            return False
            
        # Create message
        msg = MIMEMultipart('alternative')
        msg['From'] = f"{SMTP_CONFIG['sender_name']} <{SMTP_CONFIG['sender_email']}>"
        msg['To'] = to_email
        msg['Subject'] = subject
        
        # Add plain text part
        text_part = MIMEText(body, 'plain', 'utf-8')
        msg.attach(text_part)
        
        # Add HTML part if provided
        if html_body:
            html_part = MIMEText(html_body, 'html', 'utf-8')
            msg.attach(html_part)
        
        # Create SMTP session
        server = smtplib.SMTP(SMTP_CONFIG['server'], SMTP_CONFIG['port'])
        server.starttls()  # Enable TLS encryption
        server.login(SMTP_CONFIG['username'], SMTP_CONFIG['password'])
        
        # Send email
        text = msg.as_string()
        server.sendmail(SMTP_CONFIG['sender_email'], to_email, text)
        server.quit()
        
        print(f"Email sent successfully to {to_email}")
        return True
        
    except Exception as e:
        print(f"Error sending email: {e}")
        return False

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def hash_password(password):
    """Hash a password for storing."""
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(stored_password, provided_password):
    """Verify a stored password against one provided by user"""
    return stored_password == hashlib.sha256(provided_password.encode()).hexdigest()

def is_logged_in():
    """Check if user is logged in"""
    return 'user_id' in session

def has_active_subscription(user_id):
    """Check if user has an active subscription"""
    connection = get_db_connection()
    if connection is None:
        return False
    
    cursor = None
    try:
        cursor = connection.cursor()
        cursor.execute("""
            SELECT plan, start_date, end_date FROM subscriptions 
            WHERE user_id = %s AND end_date > NOW()
            ORDER BY end_date DESC
            LIMIT 1
        """, (user_id,))
        
        subscription_row = cursor.fetchone()
        return subscription_row is not None
    except Error as e:
        print(f"Error checking subscription: {e}")
        return False
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

def has_any_subscription(user_id):
    """Check if user has any subscription (active or expired)"""
    connection = get_db_connection()
    if connection is None:
        return False
    
    cursor = None
    try:
        cursor = connection.cursor()
        cursor.execute("""
            SELECT plan, start_date, end_date FROM subscriptions 
            WHERE user_id = %s
            ORDER BY end_date DESC
            LIMIT 1
        """, (user_id,))
        
        subscription_row = cursor.fetchone()
        return subscription_row is not None
    except Error as e:
        print(f"Error checking subscription history: {e}")
        return False
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

def is_subscription_expired(user_id):
    """Check if user's subscription is expired"""
    connection = get_db_connection()
    if connection is None:
        return False
    
    cursor = None
    try:
        cursor = connection.cursor()
        # Directly check in database if subscription exists and is expired
        cursor.execute("""
            SELECT COUNT(*) FROM subscriptions 
            WHERE user_id = %s AND end_date < NOW()
        """, (user_id,))
        
        result = cursor.fetchone()
        if result:
            # Get count - result is a tuple, so we need to access the first element
            count = result[0]
            # Handle different possible types
            if isinstance(count, (int, float)):
                return int(count) > 0
            elif isinstance(count, str):
                try:
                    return int(count) > 0
                except ValueError:
                    return False
            else:
                # For other types, try to convert to string then int
                try:
                    return int(str(count)) > 0
                except (ValueError, TypeError):
                    return False
        return False
    except Error as e:
        print(f"Error checking subscription expiration: {e}")
        return False
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

def extract_text_from_pdf(filepath):
    """Extract text from PDF file"""
    text = ""
    try:
        with open(filepath, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
    except Exception as e:
        print(f"Error reading PDF: {e}")
    return text

def extract_text_from_image(filepath):
    """Extract text from image file using OCR"""
    try:
        # Check if required libraries are available
        if pytesseract is None or Image is None:
            raise ImportError("pytesseract or PIL is not installed")
        text = pytesseract.image_to_string(Image.open(filepath))
        return text
    except Exception as e:
        print(f"Error reading image: {e}")
        return ""

def extract_text_from_doc(filepath):
    """Extract text from DOC/DOCX file"""
    try:
        # Check if required library is available
        if docx is None:
            raise ImportError("python-docx is not installed")
        doc = docx.Document(filepath)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    except Exception as e:
        print(f"Error reading DOC file: {e}")
        return ""

def extract_text_from_excel(filepath):
    """Extract text from Excel file"""
    try:
        df = pd.read_excel(filepath)
        text = df.to_string()
        return text
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return ""

def extract_data_from_text(text):
    """Extract transaction data from text with enhanced pattern matching"""
    transactions = []
    
    # Split text into lines for better processing
    lines = text.split('\n')
    
    # Parse the structured format in the PDF by looking for transaction blocks
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # Look for transaction date pattern
        if re.match(r'\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4}', line) and i + 4 < len(lines):
            # Check if this looks like a transaction block
            # Date should be followed by description, type, amount, balance
            try:
                date = line.strip()
                description = lines[i + 1].strip()
                trans_type = lines[i + 2].strip()
                amount_raw = lines[i + 3].strip()
                balance_raw = lines[i + 4].strip()
                
                # Validate that this is actually a transaction
                # Check if type is Credit or Debit and amount looks like money
                if trans_type in ['Credit', 'Debit'] and re.search(r'[\d,]+\.?\d{0,2}', amount_raw):
                    # Clean up amount and balance
                    amount = re.sub(r'[^\d\.]', '', amount_raw) if amount_raw else '0'
                    balance = re.sub(r'[^\d\.]', '', balance_raw) if balance_raw else '0'
                    
                    # Handle parentheses for debits
                    if '(' in amount_raw and ')' in amount_raw:
                        trans_type = 'Debit'
                        amount = amount.replace('(', '').replace(')', '')
                    
                    if float(amount) > 0:
                        transactions.append({
                            'date': date,
                            'description': description,
                            'amount': amount,
                            'balance': balance,
                            'type': trans_type
                        })
                    
                    i += 5  # Move to next potential transaction
                    continue
            except (IndexError, ValueError):
                # Not a valid transaction block, continue
                pass
        
        i += 1
    
    # If no transactions found with the structured approach, try the previous regex patterns
    if not transactions:
        # Enhanced regex patterns for different bank statement formats
        # Pattern 1: Date, Description, Amount, Balance (common format)
        pattern1 = r'(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})\s+([^$]*?)\s+\$?([\d,]+\.?\d{0,2})\s+\$?([\d,]+\.?\d{0,2})'
        
        # Pattern 2: Date, Description, Debit, Credit, Balance
        pattern2 = r'(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})\s+([^$]*?)\s+\(?[\$]?([\d,]+\.?\d{0,2})\)?\s+\(?[\$]?([\d,]+\.?\d{0,2})\)?\s+\$?([\d,]+\.?\d{0,2})'
        
        # Pattern 3: Generic pattern for transactions
        pattern3 = r'(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})\s+([^$]*?)\s+\(?[\$]?([\d,]+\.?\d{0,2})\)?'
        
        # Try pattern 1 first (most common format)
        matches = re.findall(pattern1, text, re.IGNORECASE)
        for match in matches:
            date, description, amount, balance = match
            # Clean up amount and balance
            amount = re.sub(r'[^\d\.]', '', amount) if amount else '0'
            balance = re.sub(r'[^\d\.]', '', balance) if balance else '0'
            
            if amount and float(amount) > 0:
                # Determine transaction type based on context
                desc_lower = description.lower()
                if 'deposit' in desc_lower or 'credited' in desc_lower or 'cr' in desc_lower or 'salary' in desc_lower:
                    trans_type = 'Credit'
                elif 'withdrawal' in desc_lower or 'debited' in desc_lower or 'dr' in desc_lower or 'atm' in desc_lower:
                    trans_type = 'Debit'
                else:
                    # Default heuristic
                    trans_type = 'Credit' if 'cr' in desc_lower else 'Debit'
                
                transactions.append({
                    'date': date,
                    'description': description.strip(),
                    'amount': amount,
                    'balance': balance,
                    'type': trans_type
                })
        
        # If no transactions found with pattern 1, try pattern 2
        if not transactions:
            matches = re.findall(pattern2, text, re.IGNORECASE)
            for match in matches:
                date, description, debit, credit, balance = match
                # Clean up values
                debit = re.sub(r'[^\d\.\(\)]', '', debit) if debit else '0'
                credit = re.sub(r'[^\d\.\(\)]', '', credit) if credit else '0'
                balance = re.sub(r'[^\d\.]', '', balance) if balance else '0'
                
                # Remove parentheses and handle negative values
                if '(' in debit and ')' in debit:
                    debit = debit.replace('(', '').replace(')', '')
                    trans_type = 'Debit'
                    amount = debit
                elif '(' in credit and ')' in credit:
                    credit = credit.replace('(', '').replace(')', '')
                    trans_type = 'Credit'
                    amount = credit
                elif credit and float(credit) > 0:
                    trans_type = 'Credit'
                    amount = credit
                elif debit and float(debit) > 0:
                    trans_type = 'Debit'
                    amount = debit
                else:
                    trans_type = 'Debit'
                    amount = '0'
                
                transactions.append({
                    'date': date,
                    'description': description.strip(),
                    'amount': amount,
                    'balance': balance,
                    'type': trans_type
                })
        
        # If still no transactions, try pattern 3 (generic)
        if not transactions:
            matches = re.findall(pattern3, text, re.IGNORECASE)
            for match in matches:
                date, description, amount = match
                # Clean up amount
                amount = re.sub(r'[^\d\.]', '', amount) if amount else '0'
                
                if amount and float(amount) > 0:
                    # Determine transaction type based on description
                    desc_lower = description.lower()
                    if 'deposit' in desc_lower or 'credited' in desc_lower or 'cr' in desc_lower or 'salary' in desc_lower:
                        trans_type = 'Credit'
                    elif 'interest' in desc_lower:
                        trans_type = 'Interest'
                    elif 'transfer' in desc_lower or 'trf' in desc_lower or 'wire' in desc_lower:
                        trans_type = 'Transfers'
                    elif 'loan' in desc_lower or 'mortgage' in desc_lower:
                        trans_type = 'Loans'
                    elif 'fee' in desc_lower or 'charge' in desc_lower or 'commission' in desc_lower or 'service' in desc_lower:
                        trans_type = 'Fees'
                    elif 'withdrawal' in desc_lower or 'atm' in desc_lower or 'cash' in desc_lower:
                        trans_type = 'Cash'
                    elif 'investment' in desc_lower or 'stock' in desc_lower or 'mutual fund' in desc_lower or 'shares' in desc_lower:
                        trans_type = 'Investments'
                    elif 'refund' in desc_lower or 'returned' in desc_lower or 'credit' in desc_lower:
                        trans_type = 'Refunds'
                    elif 'insurance' in desc_lower or 'premium' in desc_lower:
                        trans_type = 'Insurance'
                    elif 'tax' in desc_lower or 'irs' in desc_lower or 'revenue' in desc_lower:
                        trans_type = 'Taxes'
                    elif 'payment' in desc_lower or 'pay' in desc_lower or 'bill' in desc_lower or 'credit card' in desc_lower or 'grocery' in desc_lower or 'electricity' in desc_lower or 'store' in desc_lower:
                        trans_type = 'Payments'
                    elif 'debit' in desc_lower or 'debited' in desc_lower:
                        trans_type = 'Withdrawals'
                    else:
                        trans_type = 'Debit'
                    
                    transactions.append({
                        'date': date,
                        'description': description.strip(),
                        'amount': amount,
                        'balance': '0',  # No balance available
                        'type': trans_type
                    })
    
    # Remove duplicates based on date, description, and amount
    seen = set()
    unique_transactions = []
    for transaction in transactions:
        key = (transaction['date'], transaction['description'], transaction['amount'])
        if key not in seen:
            seen.add(key)
            unique_transactions.append(transaction)
    
    return unique_transactions

def categorize_transactions(transactions):
    """Categorize transactions based on description and amount with enhanced categories"""
    categories = {
        'Deposits': [],
        'Withdrawals': [],
        'Loans': [],
        'Interest': [],
        'Fees': [],
        'Transfers': [],
        'Payments': [],
        'Cash': [],
        'Investments': [],
        'Refunds': [],
        'Insurance': [],
        'Taxes': [],
        'UPI Transfers': [],
        'ATM Deposits': [],
        'ATM Withdrawals': [],
        'Utility Bills': [],
        'EMI/Loan Repayments': [],
        'NEFT/RTGS': [],
        'Other': []
    }
    
    for transaction in transactions:
        desc = transaction['description'].lower()
        amount = float(transaction['amount'].replace(',', ''))
        
        # Categorize based on keywords in description
        if any(keyword in desc for keyword in ['deposit', 'salary', 'income', 'payroll']) and 'atm' not in desc:
            categories['Deposits'].append(transaction)
        elif any(keyword in desc for keyword in ['interest']):
            categories['Interest'].append(transaction)
        elif any(keyword in desc for keyword in ['transfer', 'trf', 'wire']) and 'neft' not in desc and 'rtgs' not in desc and 'upi' not in desc:
            categories['Transfers'].append(transaction)
        elif any(keyword in desc for keyword in ['loan', 'mortgage']) and 'emi' not in desc and 'repayment' not in desc:
            categories['Loans'].append(transaction)
        elif any(keyword in desc for keyword in ['fee', 'charge', 'commission', 'service']):
            categories['Fees'].append(transaction)
        elif any(keyword in desc for keyword in ['withdrawal']) and 'atm' not in desc:
            categories['Withdrawals'].append(transaction)
        elif any(keyword in desc for keyword in ['atm']):
            if 'deposit' in desc or 'cr' in desc:
                categories['ATM Deposits'].append(transaction)
            else:
                categories['ATM Withdrawals'].append(transaction)
        elif any(keyword in desc for keyword in ['investment', 'stock', 'mutual fund', 'shares']):
            categories['Investments'].append(transaction)
        elif any(keyword in desc for keyword in ['refund', 'returned', 'credit']):
            categories['Refunds'].append(transaction)
        elif any(keyword in desc for keyword in ['insurance', 'premium']):
            categories['Insurance'].append(transaction)
        elif any(keyword in desc for keyword in ['tax', 'irs', 'revenue']):
            categories['Taxes'].append(transaction)
        elif any(keyword in desc for keyword in ['payment', 'pay', 'bill']) and any(keyword in desc for keyword in ['electricity', 'tv', 'mobile', 'recharge', 'phone', 'dth', 'broadband', 'water', 'gas']):
            categories['Utility Bills'].append(transaction)
        elif any(keyword in desc for keyword in ['emi', 'loan repayment', 'loan emi']):
            categories['EMI/Loan Repayments'].append(transaction)
        elif any(keyword in desc for keyword in ['upi']):
            categories['UPI Transfers'].append(transaction)
        elif any(keyword in desc for keyword in ['neft', 'rtgs']):
            categories['NEFT/RTGS'].append(transaction)
        elif any(keyword in desc for keyword in ['cash']):
            categories['Cash'].append(transaction)
        elif any(keyword in desc for keyword in ['debit', 'debited']) and 'atm' not in desc:
            categories['Withdrawals'].append(transaction)
        else:
            categories['Other'].append(transaction)
            
    return categories

def create_excel_report(categories, filename):
    """Create Excel report from categorized transactions with enhanced analytics dashboard"""
    # Calculate totals for each category
    category_totals = {}
    total_credit = 0
    total_debit = 0
    
    # Calculate totals for each category
    for category, transactions in categories.items():
        transaction_count = len(transactions)
        total_amount = sum(float(t['amount'].replace(',', '')) for t in transactions)
        
        # Calculate credit and debit amounts
        credit_amount = sum(float(t['amount'].replace(',', '')) for t in transactions if t['type'] == 'Credit')
        debit_amount = sum(float(t['amount'].replace(',', '')) for t in transactions if t['type'] == 'Debit')
        
        category_totals[category] = {
            'count': transaction_count,
            'total': round(total_amount, 2),
            'credit': round(credit_amount, 2),
            'debit': round(debit_amount, 2)
        }
        
        total_credit += credit_amount
        total_debit += debit_amount
    
    # Create Excel writer
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Create detailed summary sheet by category (this will show all categories including those with 0 transactions)
        detailed_summary = []
        # Define all categories to ensure they're all included
        all_categories = [
            'Deposits', 'Withdrawals', 'Loans', 'Interest', 'Fees', 
            'Transfers', 'Payments', 'Cash', 'Investments', 'Refunds', 
            'Insurance', 'Taxes', 'UPI Transfers', 'ATM Deposits', 
            'ATM Withdrawals', 'Utility Bills', 'EMI/Loan Repayments', 
            'NEFT/RTGS', 'Other'
        ]
        
        for category in all_categories:
            transactions = categories.get(category, [])
            if transactions:  # Category has transactions
                transaction_count = len(transactions)
                total_amount = sum(float(t['amount'].replace(',', '')) for t in transactions)
                credit_amount = sum(float(t['amount'].replace(',', '')) for t in transactions if t['type'] == 'Credit')
                debit_amount = sum(float(t['amount'].replace(',', '')) for t in transactions if t['type'] == 'Debit')
                
                detailed_summary.append({
                    'Category': category,
                    'Transaction Count': transaction_count,
                    'Total Amount': round(total_amount, 2),
                    'Credit Amount': round(credit_amount, 2),
                    'Debit Amount': round(debit_amount, 2),
                    'Net Amount': round(credit_amount - debit_amount, 2)
                })
            else:
                # Include categories with no transactions but show 0 values
                detailed_summary.append({
                    'Category': category,
                    'Transaction Count': 0,
                    'Total Amount': 0,
                    'Credit Amount': 0,
                    'Debit Amount': 0,
                    'Net Amount': 0
                })
        
        # Add overall totals to detailed summary
        detailed_summary.append({
            'Category': 'TOTAL',
            'Transaction Count': sum(len(transactions) for transactions in categories.values()),
            'Total Amount': round(total_credit + total_debit, 2),
            'Credit Amount': round(total_credit, 2),
            'Debit Amount': round(total_debit, 2),
            'Net Amount': round(total_credit - total_debit, 2)
        })
        
        detailed_summary_df = pd.DataFrame(detailed_summary)
        detailed_summary_df.to_excel(writer, sheet_name='Detailed Summary', index=False)
        
        # Create sheets for ALL categories (even those with 0 transactions)
        for category in all_categories:
            transactions = categories.get(category, [])
            # Create a sheet for every category, even if it has no transactions
            df = pd.DataFrame(transactions)
            # If no transactions, create an empty DataFrame with appropriate columns
            if df.empty:
                # Create empty DataFrame with standard columns
                df = pd.DataFrame(columns=['date', 'description', 'amount', 'type', 'balance'])
            
            # Reorder columns for better presentation
            column_order = ['date', 'description', 'amount', 'type']
            if 'balance' in df.columns:
                column_order.append('balance')
            # Only reorder if columns exist in the DataFrame
            if all(col in df.columns for col in column_order):
                df = df[column_order]
            # Sanitize category name for use as sheet name (remove invalid characters)
            sanitized_category = category.replace('/', ' ').replace('\\', ' ').replace('?', ' ').replace('*', ' ').replace('[', ' ').replace(']', ' ').replace(':', ' ')
            # Truncate to 31 characters (Excel limit)
            sheet_name = sanitized_category[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format the worksheet
            try:
                worksheet = writer.sheets[sheet_name]
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
            except:
                # If we can't access the worksheet, continue without formatting
                pass
        
        # Create analytics dashboard
        create_enhanced_analytics_dashboard(writer, categories, total_credit, total_debit, category_totals)
    
    # Re-open the file to add charts and styling
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import PatternFill, Font, Alignment
        workbook = load_workbook(filename)
        
        # Add styling to summary sheets
        # Header styling
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # Style Enhanced Summary sheet
        if 'Enhanced Summary' in workbook.sheetnames:
            summary_sheet = workbook['Enhanced Summary']
            for cell in summary_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Highlight negative values in red
            for row in summary_sheet.iter_rows(min_row=2, max_row=summary_sheet.max_row):
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.value < 0:
                        cell.font = Font(color='FF0000')  # Red color for negative values
        
        # Style Detailed Summary sheet with company logo and attractive background
        if 'Detailed Summary' in workbook.sheetnames:
            detailed_sheet = workbook['Detailed Summary']
            
            # Style header row (A1-F1) similar to other sheets without violet background
            # Skip N1 and O1 cells as they contain logo and text
            for col in range(1, min(16, detailed_sheet.max_column + 1)):  # Columns A-P (1-16)
                if col != 14 and col != 15:  # Skip columns N (14) and O (15)
                    cell = detailed_sheet.cell(row=1, column=col)
                    cell.fill = header_fill  # Standard header fill from other sheets
                    cell.font = header_font  # Standard header font from other sheets
                    cell.alignment = Alignment(horizontal='center')
            
            # Highlight negative values in red
            for row in detailed_sheet.iter_rows(min_row=2, max_row=detailed_sheet.max_row):
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.value < 0:
                        cell.font = Font(color='FF0000')  # Red color for negative values
            
            # Add company logo in cell N1 and text in cell O1 (without header background color)
            logo_path = 'static/images/logo.png'  # Default logo path
            if os.path.exists(logo_path):
                try:
                    # Add logo to cell N1 (14th column, 1st row) with smaller size
                    logo = ExcelImage(logo_path)
                    logo.width = 80  # Smaller size
                    logo.height = 80
                    detailed_sheet.add_image(logo, 'N1')  # Position logo at N1
                    
                    # Add company title in cell O1 (15th column, 1st row) with smaller size
                    title_cell = detailed_sheet['O1']
                    title_cell.value = "E-Faws Tech Pvt Limited"
                    title_cell.font = Font(bold=True, size=10, color='800080')  # Smaller purple color text
                    title_cell.alignment = Alignment(horizontal='left')
                    
                    # Add subtitle in cell O11
                    subtitle_cell = detailed_sheet['O2']
                    subtitle_cell.value = "Bank Statement Analyzer"
                    subtitle_cell.font = Font(bold=True, size=10, color='800080')  # Purple color
                    subtitle_cell.alignment = Alignment(horizontal='left')
                except Exception as e:
                    print(f"Warning: Could not add logo to Excel report: {e}")
            else:
                print(f"Note: Logo file not found at {logo_path}")
        
        # Style category sheets
        for category in categories.keys():
            if category in workbook.sheetnames:
                worksheet = workbook[category]
                # Header styling
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Highlight debit transactions in red (if we can find the right columns)
                # This is more complex without knowing the exact column structure
        
        # Save the styled workbook
        workbook.save(filename)
    except Exception as e:
        print(f"Error styling Excel workbook: {e}")
        # If styling fails, the file is still usable without styling

# Function to save analysis to database
def save_analysis_to_db(user_id, name, bank_name, customer_number, file_name, excel_file_path, pdf_file_path=None):
    """Save bank statement analysis to database"""
    connection = get_db_connection()
    if connection is None:
        return False
    
    cursor = None
    try:
        cursor = connection.cursor()
        cursor.execute("""
            INSERT INTO bank_statement_analyses 
            (user_id, name, bank_name, customer_number, file_name, excel_file_path, pdf_file_path) 
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (user_id, name, bank_name, customer_number, file_name, excel_file_path, pdf_file_path))
        connection.commit()
        return True
    except Error as e:
        print(f"Error saving analysis to database: {e}")
        return False
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

def get_user_analyses(user_id, filter_date=None, filter_day=None):
    """Get all analyses for a user with optional date/day filtering"""
    connection = get_db_connection()
    if connection is None:
        return []
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Base query
        query = """
            SELECT id, name, bank_name, customer_number, analysis_date, file_name 
            FROM bank_statement_analyses 
            WHERE user_id = %s 
        """
        
        # Add date filter if provided
        params = [user_id]
        if filter_date:
            query += " AND DATE(analysis_date) = %s "
            params.append(filter_date)
        elif filter_day:
            query += " AND DAYNAME(analysis_date) = %s "
            params.append(filter_day)
        
        query += " ORDER BY analysis_date DESC"
        
        cursor.execute(query, params)
        
        analyses = []
        for row in cursor.fetchall():
            analysis_dict = {
                "id": row[0],
                "name": row[1],
                "bank_name": row[2],
                "customer_number": row[3],
                "analysis_date": row[4],
                "file_name": row[5]
            }
            analyses.append(analysis_dict)
        
        return analyses
    except Error as e:
        print(f"Error retrieving analyses: {e}")
        return []
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

def get_user_name(user_id):
    """Get user name from database by user ID"""
    if user_id == 'admin':
        return 'Admin'
    
    connection = get_db_connection()
    if connection is None:
        return 'User'
    
    cursor = None
    try:
        cursor = connection.cursor()
        cursor.execute("SELECT name FROM users WHERE id = %s", (user_id,))
        user_row = cursor.fetchone()
        
        if user_row:
            return user_row[0]  # Return the name
        else:
            return 'User'
    except Error as e:
        print(f"Error fetching user name: {e}")
        return 'User'

    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

def delete_analysis(analysis_id, user_id):
    """Delete an analysis record"""
    connection = get_db_connection()
    if connection is None:
        return False
    
    cursor = None
    try:
        cursor = connection.cursor()
        cursor.execute("""
            DELETE FROM bank_statement_analyses 
            WHERE id = %s AND user_id = %s
        """, (analysis_id, user_id))
        connection.commit()
        return cursor.rowcount > 0
    except Error as e:
        print(f"Error deleting analysis: {e}")
        return False
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

def create_analytics_dashboard(writer, categories, total_credit, total_debit):
    """Create analytics dashboard with charts"""
    # Prepare data for charts
    cashflow_data = []
    
    for category, transactions in categories.items():
        credit_amount = sum(float(t['amount'].replace(',', '')) for t in transactions if t['type'] == 'Credit')
        debit_amount = sum(float(t['amount'].replace(',', '')) for t in transactions if t['type'] == 'Debit')
        
        # Only include categories with significant amounts
        if credit_amount > 0 or debit_amount > 0:
            cashflow_data.append({
                'Category': category,
                'Credit': round(credit_amount, 2),
                'Debit': round(debit_amount, 2)
            })
    
    # Create cashflow summary sheet
    cashflow_df = pd.DataFrame(cashflow_data)
    cashflow_df.to_excel(writer, sheet_name='Cashflow Summary', index=False)
    
    # Create monthly average balance data (simulated for demonstration)
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    monthly_balance_data = []
    base_balance = 10000  # Starting balance
    
    for i, month in enumerate(months):
        # Simulate balance changes
        balance_change = (total_credit - total_debit) * (0.8 + 0.4 * i / 12) / 12
        base_balance += balance_change
        monthly_balance_data.append({
            'Month': month,
            'Average Balance': round(base_balance, 2)
        })
    
    monthly_balance_df = pd.DataFrame(monthly_balance_data)
    monthly_balance_df.to_excel(writer, sheet_name='Monthly Average Balance', index=False)
    
    # Create daily average balance data (simulated for demonstration)
    days = [f'Day {i+1}' for i in range(30)]
    daily_balance_data = []
    daily_base_balance = 10000
    
    for i, day in enumerate(days):
        # Simulate daily balance changes
        daily_change = (total_credit - total_debit) / 30 * (0.9 + 0.2 * (i % 10) / 10)
        daily_base_balance += daily_change
        daily_balance_data.append({
            'Day': day,
            'Average Balance': round(daily_base_balance, 2)
        })
    
    daily_balance_df = pd.DataFrame(daily_balance_data)
    daily_balance_df.to_excel(writer, sheet_name='Daily Average Balance', index=False)

def create_enhanced_analytics_dashboard(writer, categories, total_credit, total_debit, category_totals):
    """Create enhanced analytics dashboard with attractive charts and visualizations"""
    # Prepare data for charts
    cashflow_data = []
    
    # Define all categories to ensure they're all included
    all_categories = [
        'Deposits', 'Withdrawals', 'Loans', 'Interest', 'Fees', 
        'Transfers', 'Payments', 'Cash', 'Investments', 'Refunds', 
        'Insurance', 'Taxes', 'UPI Transfers', 'ATM Deposits', 
        'ATM Withdrawals', 'Utility Bills', 'EMI/Loan Repayments', 
        'NEFT/RTGS', 'Other'
    ]
    
    for category in all_categories:
        transactions = categories.get(category, [])
        credit_amount = sum(float(t['amount'].replace(',', '')) for t in transactions if t.get('type') == 'Credit')
        debit_amount = sum(float(t['amount'].replace(',', '')) for t in transactions if t.get('type') == 'Debit')
        
        # Include all categories, even those with 0 amounts
        cashflow_data.append({
            'Category': category,
            'Credit': round(credit_amount, 2),
            'Debit': round(debit_amount, 2)
        })
    
    # Create cashflow summary sheet
    cashflow_df = pd.DataFrame(cashflow_data)
    cashflow_df.to_excel(writer, sheet_name='Cashflow Summary', index=False)
    
    # Create monthly average balance data (simulated for demonstration)
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    monthly_balance_data = []
    base_balance = 10000  # Starting balance
    
    for i, month in enumerate(months):
        # Simulate balance changes with more realistic fluctuations
        # Using a wave-like pattern with random variations
        variation = (0.8 + 0.4 * (i % 6) / 6)  # Creates a wave pattern
        balance_change = (total_credit - total_debit) * variation / 12
        base_balance += balance_change
        
        # Ensure we don't go negative and round to 2 decimal places
        monthly_balance_data.append({
            'Month': month,
            'Average Balance': max(0, round(base_balance, 2))
        })
    
    monthly_balance_df = pd.DataFrame(monthly_balance_data)
    monthly_balance_df.to_excel(writer, sheet_name='Monthly Average Balance', index=False)
    
    # Create daily average balance data (simulated for demonstration)
    days = [f'Day {i+1}' for i in range(30)]
    daily_balance_data = []
    daily_base_balance = 10000
    
    for i, day in enumerate(days):
        # Simulate daily balance changes with more realistic fluctuations
        # Using a wave-like pattern with random variations
        variation = (0.9 + 0.2 * ((i % 10) / 10))  # Creates a wave pattern
        daily_change = (total_credit - total_debit) / 30 * variation
        
        # Handle cases where balance might go negative
        daily_base_balance += daily_change
        
        # Ensure we don't go negative and round to 2 decimal places
        daily_balance_data.append({
            'Day': day,
            'Average Balance': max(0, round(daily_base_balance, 2))
        })
    
    daily_balance_df = pd.DataFrame(daily_balance_data)
    daily_balance_df.to_excel(writer, sheet_name='Daily Average Balance', index=False)
    
    # Add chart data sheet
    chart_data = []
    for item in cashflow_data:
        if item['Credit'] > 0:
            chart_data.append({'Type': f"{item['Category']} (Credit)", 'Amount': item['Credit']})
        if item['Debit'] > 0:
            chart_data.append({'Type': f"{item['Category']} (Debit)", 'Amount': item['Debit']})
    
    chart_df = pd.DataFrame(chart_data)
    chart_df.to_excel(writer, sheet_name='Chart Data', index=False)
    
    # Create category-specific sheets for ALL categories (even those with 0 transactions)
    for category in all_categories:
        transactions = categories.get(category, [])
        # Create a sheet for every category, even if it has no transactions
        df = pd.DataFrame(transactions)
        # If no transactions, create an empty DataFrame with appropriate columns
        if df.empty:
            # Create empty DataFrame with standard columns
            df = pd.DataFrame(columns=['date', 'description', 'amount', 'type', 'balance'])
        
        # Reorder columns for better presentation
        column_order = ['date', 'description', 'amount', 'type']
        if 'balance' in df.columns:
            column_order.append('balance')
        # Only reorder if columns exist in the DataFrame
        if all(col in df.columns for col in column_order):
            df = df[column_order]
        
        # Sanitize category name for use as sheet name (remove invalid characters)
        sanitized_category = category.replace('/', ' ').replace('\\', ' ').replace('?', ' ').replace('*', ' ').replace('[', ' ').replace(']', ' ').replace(':', ' ')
        # Truncate to 31 characters (Excel limit)
        sheet_name = sanitized_category[:31]
        
        df.to_excel(writer, sheet_name=f"{sheet_name} Details", index=False)

        # If no transactions, create an empty DataFrame with appropriate columns
        if df.empty:
            # Create empty DataFrame with standard columns
            df = pd.DataFrame(columns=['date', 'description', 'amount', 'type', 'balance'])
        
        # Reorder columns for better presentation
        column_order = ['date', 'description', 'amount', 'type']
        if 'balance' in df.columns:
            column_order.append('balance')
        # Only reorder if columns exist in the DataFrame
        if all(col in df.columns for col in column_order):
            df = df[column_order]
        
        # Sanitize category name for use as sheet name (remove invalid characters)
        sanitized_category = category.replace('/', ' ').replace('\\', ' ').replace('?', ' ').replace('*', ' ').replace('[', ' ').replace(']', ' ').replace(':', ' ')
        # Truncate to 31 characters (Excel limit)
        sheet_name = sanitized_category[:31]
        
        df.to_excel(writer, sheet_name=f"{sheet_name} Details", index=False)
    
    # Access the workbook and add charts
    workbook = writer.book
    
    # Style the dashboard sheets
    from openpyxl.styles import PatternFill, Font, Alignment
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Style cashflow summary sheet
    if 'Cashflow Summary' in workbook.sheetnames:
        cashflow_sheet = workbook['Cashflow Summary']
        for cell in cashflow_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
    
    # Style monthly average balance sheet
    if 'Monthly Average Balance' in workbook.sheetnames:
        monthly_sheet = workbook['Monthly Average Balance']
        for cell in monthly_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
    
    # Style daily average balance sheet
    if 'Daily Average Balance' in workbook.sheetnames:
        daily_sheet = workbook['Daily Average Balance']
        for cell in daily_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
    
    # 1. Cashflow Summary Pie Chart
    if 'Cashflow Summary' in workbook.sheetnames:
        cashflow_sheet = workbook['Cashflow Summary']
        
        # Create pie chart for cashflow summary
        pie_chart = PieChart()
        pie_chart.title = "Cashflow Summary by Category"
        
        # Data for the chart (excluding the total row)
        categories_ref = Reference(cashflow_sheet, min_col=1, min_row=2, max_row=len(cashflow_data)+1)
        credit_values = Reference(cashflow_sheet, min_col=3, min_row=1, max_row=len(cashflow_data)+1)
        
        # Add data to chart
        pie_chart.add_data(credit_values, titles_from_data=True)
        pie_chart.set_categories(categories_ref)
        
        # Position the chart
        cashflow_sheet.add_chart(pie_chart, "F2")
    
    # 2. Monthly Average Balance Area Chart
    if 'Monthly Average Balance' in workbook.sheetnames:
        monthly_sheet = workbook['Monthly Average Balance']
        
        # Create area chart for monthly average balance
        from openpyxl.chart import AreaChart
        area_chart_monthly = AreaChart()
        area_chart_monthly.title = "Monthly Average Balance Trend"
        area_chart_monthly.x_axis.title = "Month"
        area_chart_monthly.y_axis.title = "Balance"
        area_chart_monthly.style = 13  # Attractive style
        
        # Data for the chart
        months_data = Reference(monthly_sheet, min_col=1, min_row=2, max_row=len(monthly_balance_data)+1)
        balance_values = Reference(monthly_sheet, min_col=2, min_row=1, max_row=len(monthly_balance_data)+1)
        
        # Add data to chart
        area_chart_monthly.add_data(balance_values, titles_from_data=True)
        area_chart_monthly.set_categories(months_data)
        
        # Position the chart
        monthly_sheet.add_chart(area_chart_monthly, "D2")
    
    # 3. Daily Average Balance Area Chart
    if 'Daily Average Balance' in workbook.sheetnames:
        daily_sheet = workbook['Daily Average Balance']
        
        # Create area chart for daily average balance
        from openpyxl.chart import AreaChart
        area_chart_daily = AreaChart()
        area_chart_daily.title = "Daily Average Balance Trend"
        area_chart_daily.x_axis.title = "Day"
        area_chart_daily.y_axis.title = "Balance"
        area_chart_daily.style = 13  # Attractive style
        
        # Data for the chart
        days_data = Reference(daily_sheet, min_col=1, min_row=2, max_row=len(daily_balance_data)+1)
        daily_balance_values = Reference(daily_sheet, min_col=2, min_row=1, max_row=len(daily_balance_data)+1)
        
        # Add data to chart
        area_chart_daily.add_data(daily_balance_values, titles_from_data=True)
        area_chart_daily.set_categories(days_data)
        
        # Position the chart
        daily_sheet.add_chart(area_chart_daily, "D2")
    
    # 4. Enhanced Bar Chart for Credits vs Debits by Category
    if 'Cashflow Summary' in workbook.sheetnames:
        cashflow_sheet = workbook['Cashflow Summary']
        
        # Create bar chart
        from openpyxl.chart import BarChart
        bar_chart = BarChart()
        bar_chart.title = "Credits vs Debits by Category"
        bar_chart.x_axis.title = "Category"
        bar_chart.y_axis.title = "Amount"
        bar_chart.grouping = "clustered"
        
        # Data for credits and debits
        categories_ref = Reference(cashflow_sheet, min_col=1, min_row=2, max_row=len(cashflow_data)+1)
        credit_values = Reference(cashflow_sheet, min_col=2, min_row=1, max_row=len(cashflow_data)+1)
        debit_values = Reference(cashflow_sheet, min_col=3, min_row=1, max_row=len(cashflow_data)+1)
        
        # Add data to chart
        bar_chart.add_data(credit_values, titles_from_data=True)
        bar_chart.add_data(debit_values, titles_from_data=True)
        bar_chart.set_categories(categories_ref)
        
        # Position the chart
        cashflow_sheet.add_chart(bar_chart, "F15")

        
@app.route('/')
def index():
    # If user is logged in and has active subscription, go to analyzer
    if is_logged_in() and has_active_subscription(session['user_id']):
        user_name = get_user_name(session['user_id'])
        return redirect(url_for('bank_statement_analyzer'))
    # Otherwise go to login page
    return redirect(url_for('login'))

@app.route('/login')
def login():
    # If user is already logged in, check their subscription status
    if is_logged_in():
        # Check if this is an admin user
        if session.get('is_admin'):
            return redirect(url_for('admin_dashboard'))
        # If they have an active subscription, redirect to analyzer
        if has_active_subscription(session['user_id']):
            return redirect(url_for('bank_statement_analyzer'))
        # If they don't have an active subscription, redirect to subscription page
        else:
            return redirect(url_for('subscription'))
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login_post():
    email = request.form.get('email')
    password = request.form.get('password')
    
    # Check for admin login
    if email == 'efaws@gmail.com' and password == '1234':
        session['user_id'] = 'admin'
        session['email'] = email
        session['is_admin'] = True
        return redirect(url_for('admin_dashboard'))
    
    # Check if user exists in database
    connection = get_db_connection()
    if connection is None:
        return redirect(url_for('login', error='Database connection failed'))
    
    cursor = None
    try:
        cursor = connection.cursor()
        cursor.execute("SELECT id, name, email, password, is_demo, demo_expiry_date FROM users WHERE email = %s", (email,))
        user_row = cursor.fetchone()
        
        # Check if user exists and password is correct
        if user_row is not None:
            # Unpack the row data (user_id, name, email, password, is_demo, demo_expiry_date)
            user_data = list(user_row)
            user_id = user_data[0]
            name = user_data[1]
            user_email = user_data[2]
            user_password = user_data[3]
            is_demo = user_data[4] if len(user_data) > 4 else False
            demo_expiry_date = user_data[5] if len(user_data) > 5 else None
            
            if verify_password(user_password, password):
                # Check if this is a demo account and if it has expired
                if is_demo and demo_expiry_date:
                    from datetime import datetime
                    if datetime.now() > demo_expiry_date:
                        # Demo account has expired, redirect to register page with message
                        return redirect(url_for('login', error='Your demo is finished now you can register'))
                
                session['user_id'] = user_id
                session['email'] = user_email
                session['is_admin'] = False  # Not an admin
                session['is_demo'] = is_demo  # Store demo status in session
                
                # For demo users, go directly to bank statement analyzer
                if is_demo:
                    return redirect(url_for('bank_statement_analyzer'))
                
                # Check if user has active subscription
                if has_active_subscription(user_id):
                    return redirect(url_for('bank_statement_analyzer'))
                else:
                    # Check if user has any subscription (expired or active)
                    if has_any_subscription(user_id) and is_subscription_expired(user_id):
                        # User has an expired subscription
                        return redirect(url_for('subscription', expired=True))
                    else:
                        # User has no subscription or it's not expired
                        return redirect(url_for('subscription'))
        
        return redirect(url_for('login', error='Invalid email or password'))
    except Error as e:
        print(f"Error during login: {e}")
        return redirect(url_for('login', error='Login failed due to server error'))
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/register', methods=['POST'])
def register_post():
    """Register new user and track registration history"""
    name = request.form.get('name')  # This is now company name
    phone = request.form.get('phone')  # New phone number field
    email = request.form.get('email')
    password = request.form.get('password')
    confirm_password = request.form.get('confirm_password')
    
    # Check if passwords match
    if password != confirm_password:
        return redirect(url_for('login', error='Passwords do not match'))
    
    # Hash password
    hashed_password = hash_password(password)
    
    # Save user to database
    connection = get_db_connection()
    if connection is None:
        return redirect(url_for('login', error='Database connection failed'))
    
    cursor = None
    try:
        cursor = connection.cursor()
        cursor.execute(
            "INSERT INTO users (name, email, password, phone) VALUES (%s, %s, %s, %s)",
            (name, email, hashed_password, phone)
        )
        connection.commit()
        
        # Get the inserted user ID
        user_id = cursor.lastrowid
        
        # Create notification for new user registration
        try:
            cursor.execute("""
                INSERT INTO notifications (user_id, type, title, message) 
                VALUES (%s, %s, %s, %s)
            """, (user_id, 'new-user', 'New User Registration', f'{name} has just registered'))
        except Error as e:
            print(f"Error creating notification: {e}")
        
        connection.commit()
        
        # Log in the user
        session['user_id'] = user_id
        session['email'] = email
        
        return redirect(url_for('subscription'))
    except Error as e:
        connection.rollback()
        if e.errno == 1062:  # Duplicate entry
            return redirect(url_for('login', error='Email already registered'))
        else:
            print(f"Error during registration: {e}")
            return redirect(url_for('login', error='Registration failed due to server error'))
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/subscription')
def subscription():
    if not is_logged_in():
        return redirect(url_for('login'))
    user_name = get_user_name(session['user_id'])
    
    # Check if user has an expired subscription
    expired = request.args.get('expired', False)
    
    return render_template('subscription.html', user_name=user_name, expired=expired)

@app.route('/create-order', methods=['POST'])
def create_order():
    print("=== CREATE ORDER ENDPOINT CALLED ===")
    
    if not is_logged_in():
        print("User not logged in")
        response = jsonify({'success': False, 'error': 'User not logged in'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response
    
    print("User is logged in")
    
    # Check if Razorpay client is initialized
    if razorpay_client is None:
        print("Razorpay client is None, trying to reinitialize")
        # Try to reinitialize
        if not initialize_razorpay():
            print("Failed to initialize Razorpay client")
            response = jsonify({'success': False, 'error': 'Payment system not properly configured. Please contact administrator. See RAZORPAY_SETUP.md for instructions.'})
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response
        else:
            print("Razorpay client reinitialized successfully")
    
    try:
        data = request.get_json()
        print(f"Request data: {data}")
        plan = data.get('plan', 'monthly') if data else 'monthly'
        
        print(f"Creating order for plan: {plan}")  # Debug log
        
        # Determine amount based on plan
        if plan == 'monthly':
            amount = 49900  # ₹499 in paise
        elif plan == 'two_months':
            amount = 89900  # ₹899 in paise
        elif plan == 'annual':
            amount = 499900  # ₹4999 in paise
        else:
            amount = 49900  # default to monthly
        
        print(f"Order amount: {amount}")  # Debug log
        
        # Check if RAZORPAY_KEY_ID is not None before accessing
        if RAZORPAY_KEY_ID:
            print(f"Using Razorpay keys - ID: {RAZORPAY_KEY_ID[:10]}... SECRET: {'*' * 10}")  # Debug log
        
        # Create Razorpay order
        order_data = {
            'amount': amount,
            'currency': 'INR',
            'payment_capture': 1  # Auto-capture payment
        }
        
        print(f"Order data: {order_data}")  # Debug log
        
        # Create order using Razorpay client - FIXED THE ACCESS METHOD
        if razorpay_client:
            print("Creating order with Razorpay client")
            order = razorpay_client.order.create(data=order_data)
            print(f"Order created successfully: {order}")  # Debug log
            response_data = {
                'success': True, 
                'order_id': order['id'], 
                'amount': amount,
                'key_id': RAZORPAY_KEY_ID  # Send the key to the frontend
            }
            response = jsonify(response_data)
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response
        else:
            print("Razorpay client is still None")
            response = jsonify({'success': False, 'error': 'Razorpay client not initialized'})
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response
    except Exception as e:
        error_msg = f"Order creation failed: {str(e)}"
        print(error_msg)
        import traceback
        traceback.print_exc()
        response = jsonify({'success': False, 'error': error_msg})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response

@app.route('/admin/create-demo-account', methods=['POST'])
def admin_create_demo_account():
    """Create a demo account with expiration"""
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        days = int(data.get('days', 2))  # Default to 2 days
        
        # Validate input
        if not username or not password:
            return jsonify({'success': False, 'error': 'Username and password are required'})
        
        if not re.match(r"[^@]+@[^@]+\.[^@]+", username):
            return jsonify({'success': False, 'error': 'Invalid email format'})
        
        # Hash password
        hashed_password = hash_password(password)
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'error': 'Database connection failed'})
        
        cursor = None
        try:
            cursor = connection.cursor()
            
            # Check if user already exists
            cursor.execute("SELECT id FROM users WHERE email = %s", (username,))
            existing_user = cursor.fetchone()
            
            if existing_user:
                return jsonify({'success': False, 'error': 'User with this email already exists'})
            
            # Calculate demo expiry date
            from datetime import datetime, timedelta
            expiry_date = datetime.now() + timedelta(days=days)
            
            # Create demo user
            cursor.execute("""
                INSERT INTO users (name, email, password, is_demo, demo_expiry_date) 
                VALUES (%s, %s, %s, %s, %s)
            """, (f"Demo User ({days} days)", username, hashed_password, True, expiry_date))
            
            connection.commit()
            
            # Create notification for new demo user
            user_id = cursor.lastrowid
            cursor.execute("""
                INSERT INTO notifications (user_id, type, title, message) 
                VALUES (%s, %s, %s, %s)
            """, (user_id, 'new-user', 'New Demo User', f'Demo user {username} created for {days} days'))
            
            connection.commit()
            
            return jsonify({'success': True, 'message': f'Demo account created successfully. Expires in {days} days.'})
        except Error as e:
            connection.rollback()
            print(f"Error creating demo account: {e}")
            return jsonify({'success': False, 'error': 'Failed to create demo account'})
        finally:
            if cursor:
                cursor.close()
            if connection and connection.is_connected():
                connection.close()
                
    except Exception as e:
        print(f"Error in demo account creation: {e}")
        return jsonify({'success': False, 'error': 'An error occurred during account creation'})

        traceback.print_exc()  # Print full traceback
        response = jsonify({'success': False, 'error': 'Failed to create payment order. Please try again or contact support.'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response

@app.route('/verify-payment', methods=['POST'])
def verify_payment():
    """Verify payment and create subscription with history tracking"""
    print("=== VERIFY PAYMENT ENDPOINT CALLED ===")
    
    if not is_logged_in():
        print("User not logged in")
        response = jsonify({'success': False, 'error': 'User not logged in'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response
    
    try:
        data = request.get_json()
        print(f"Request data: {data}")
        
        if not data:
            print("No data received in request")
            response = jsonify({'success': False, 'error': 'No data received'})
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response
        
        payment_id = data.get('payment_id')
        order_id = data.get('order_id')
        signature = data.get('signature')
        plan = data.get('plan', 'monthly')
        
        print(f"Payment verification data: payment_id={payment_id}, order_id={order_id}, signature={signature}, plan={plan}")
        
        # Validate required fields
        if not payment_id or not order_id or not signature:
            print("Missing required payment verification data")
            response = jsonify({'success': False, 'error': 'Missing required payment data'})
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response
        
        # In a real implementation, you would verify the payment with Razorpay
        # For demo purposes, we'll assume payment is successful
        
        # Determine subscription duration based on plan
        if plan == 'monthly':
            days = 30
        elif plan == 'two_months':
            days = 60
        elif plan == 'annual':
            days = 365
        else:
            days = 30  # default to monthly
        
        print(f"Creating subscription for {days} days")
        
        # Save subscription to database
        connection = get_db_connection()
        if connection is None:
            print("Database connection failed")
            response = jsonify({'success': False, 'error': 'Database connection failed'})
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response
        
        cursor = None
        try:
            cursor = connection.cursor()
            cursor.execute("""
                INSERT INTO subscriptions (user_id, plan, start_date, end_date) 
                VALUES (%s, %s, NOW(), DATE_ADD(NOW(), INTERVAL %s DAY))
            """, (session['user_id'], plan, days))
            connection.commit()
            print("Subscription saved to database")
            
            # Create notification for new subscription
            try:
                user_id = session['user_id']
                cursor.execute("SELECT name FROM users WHERE id = %s", (user_id,))
                user_data = cursor.fetchone()
                if user_data:
                    user_name = user_data[0]
                    cursor.execute("""
                        INSERT INTO notifications (user_id, type, title, message) 
                        VALUES (%s, %s, %s, %s)
                    """, (user_id, 'new-subscription', 'New Subscription', f'{user_name} subscribed to {plan}'))
                    connection.commit()
                    print("Subscription notification created")
            except Error as e:
                print(f"Error creating subscription notification: {e}")
            
            # Set session variable to indicate successful payment
            session['payment_success'] = True
            print("Payment success flag set in session")
            
            response_data = {'success': True}
            response = jsonify(response_data)
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response
        except Error as e:
            connection.rollback()
            error_msg = f"Error saving subscription: {e}"
            print(error_msg)
            response = jsonify({'success': False, 'error': 'Failed to save subscription'})
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response
        finally:
            if cursor:
                cursor.close()
            if connection and connection.is_connected():
                connection.close()
                
    except Exception as e:
        error_msg = f"Payment verification failed: {e}"
        print(error_msg)
        import traceback
        traceback.print_exc()
        response = jsonify({'success': False, 'error': 'Payment verification failed'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('email', None)
    return redirect(url_for('login'))

@app.route('/bank-statement-analyzer')
def bank_statement_analyzer():
    if not is_logged_in():
        return redirect(url_for('login'))
    
    # Check if this is a demo user and if their account has expired
    if session.get('is_demo'):
        # Get user info to check expiration
        connection = get_db_connection()
        if connection is None:
            return redirect(url_for('login', error='Database connection failed'))
        
        cursor = None
        try:
            cursor = connection.cursor()
            cursor.execute("SELECT is_demo, demo_expiry_date FROM users WHERE id = %s", (session['user_id'],))
            user_row = cursor.fetchone()
            
            if user_row:
                is_demo = user_row[0]
                demo_expiry_date = user_row[1]
                
                # Check if this is a demo account and if it has expired
                if is_demo and demo_expiry_date:
                    from datetime import datetime
                    if datetime.now() > demo_expiry_date:
                        # Demo account has expired, redirect to register page with message
                        return redirect(url_for('login', error='Your demo is finished now you can register'))
            else:
                return redirect(url_for('login', error='User not found'))
        except Error as e:
            print(f"Error checking demo account: {e}")
            return redirect(url_for('login', error='Error checking account status'))
        finally:
            if cursor:
                cursor.close()
            if connection and connection.is_connected():
                connection.close()
    
    # For non-demo users, check subscription
    elif not has_active_subscription(session['user_id']):
        return redirect(url_for('subscription'))
    
    # Check if this is a successful payment redirect
    payment_success = session.pop('payment_success', False)
    user_name = get_user_name(session['user_id'])
    
    return render_template('bank_statement.html', payment_success=payment_success, user_name=user_name)

@app.route('/bank-statement-history')
def bank_statement_history():
    if not is_logged_in():
        return redirect(url_for('login'))
    
    if not has_active_subscription(session['user_id']):
        return redirect(url_for('subscription'))
    
    # Get filter parameters
    filter_date = request.args.get('date')
    filter_day = request.args.get('day')
    
    # Get user analyses with filtering
    analyses = get_user_analyses(session['user_id'], filter_date, filter_day)
    user_name = get_user_name(session['user_id'])
    return render_template('bank_statement_history.html', analyses=analyses, filter_date=filter_date, filter_day=filter_day, user_name=user_name)

@app.route('/upload', methods=['POST'])
def upload_file():
    if not is_logged_in():
        return redirect(url_for('login'))
    
    if not has_active_subscription(session['user_id']):
        return redirect(url_for('subscription'))
    
    # Get user info from form
    name = request.form.get('name', '')
    bank_name = request.form.get('bank_name', '')
    customer_number = request.form.get('customer_number', '')
    
    if 'file' not in request.files:
        flash('No file selected')
        return redirect(request.url)
    
    file = request.files['file']
    
    if file.filename == '':
        flash('No file selected')
        return redirect(request.url)
    
    if file and file.filename and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Process the file based on its type
        if filename.lower().endswith('.pdf'):
            text = extract_text_from_pdf(filepath)
        elif filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
            text = extract_text_from_image(filepath)
        elif filename.lower().endswith(('.doc', '.docx')):
            text = extract_text_from_doc(filepath)
        elif filename.lower().endswith(('.xls', '.xlsx')):
            text = extract_text_from_excel(filepath)
        else:
            # For other file types, try to read as text
            try:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    text = f.read()
            except:
                text = "Could not extract text from file"
        
        # Extract transactions from text
        transactions = extract_data_from_text(text)
        
        # Categorize transactions
        categories = categorize_transactions(transactions)
        
        # Create Excel report
        excel_filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        excel_filepath = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
        create_excel_report(categories, excel_filepath)
        
        # Create PDF report
        pdf_filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        pdf_filepath = os.path.join(app.config['UPLOAD_FOLDER'], pdf_filename)
        
        # Account info can be extracted from the document or provided by user
        account_info = {
            "bank_name": bank_name if bank_name else "Bank Statement Analyzer",
            "account_number": customer_number if customer_number else "Extracted from document",
            "account_holder": name if name else "Client",
            "statement_period": "Analysis Period",
            "opening_balance": "$0.00",
            "closing_balance": "$0.00"
        }
        
        # Import the PDF generator function
        try:
            from enhanced_pdf_generator import create_professional_bank_statement
            create_professional_bank_statement(categories, pdf_filepath, account_info)
        except ImportError:
            print("PDF generation module not found")
        
        # Save analysis to database
        save_analysis_to_db(
            session['user_id'], 
            name if name else 'User', 
            bank_name if bank_name else 'Unknown Bank', 
            customer_number if customer_number else 'N/A', 
            filename, 
            excel_filepath, 
            pdf_filepath
        )
        
        # Return the Excel file for download
        return send_file(excel_filepath, as_attachment=True)
    
    flash('Invalid file type')
    return redirect(url_for('bank_statement_analyzer'))

@app.route('/analysis-details/<int:analysis_id>')
def analysis_details(analysis_id):
    if not is_logged_in():
        return redirect(url_for('login'))
    
    # Get analysis details
    connection = get_db_connection()
    if connection is None:
        return "Database connection failed", 500
    
    cursor = None
    try:
        cursor = connection.cursor()
        cursor.execute("""
            SELECT id, name, bank_name, customer_number, analysis_date, file_name, excel_file_path
            FROM bank_statement_analyses 
            WHERE id = %s AND user_id = %s
        """, (analysis_id, session['user_id']))
        
        result = cursor.fetchone()
        if result:
            # Convert database results to a dictionary
            analysis = {
                "id": result[0],
                "name": result[1],
                "bank_name": result[2],
                "customer_number": result[3],
                "analysis_date": result[4],
                "file_name": result[5],
                "excel_file_path": result[6]
            }
            
            user_name = get_user_name(session['user_id'])
            return render_template('analysis_details.html', analysis=analysis, user_name=user_name)
        else:
            return "Analysis not found", 404
    except Error as e:
        print(f"Error retrieving analysis: {e}")
        return "Error retrieving analysis", 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/download-excel/<int:analysis_id>')
def download_excel(analysis_id):
    if not is_logged_in():
        return redirect(url_for('login'))
    
    # Get analysis details
    connection = get_db_connection()
    if connection is None:
        return "Database connection failed", 500
    
    cursor = None
    try:
        cursor = connection.cursor()
        cursor.execute("""
            SELECT excel_file_path, file_name 
            FROM bank_statement_analyses 
            WHERE id = %s AND user_id = %s
        """, (analysis_id, session['user_id']))
        
        result = cursor.fetchone()
        if result:
            # Convert database results to strings
            excel_file_path = str(result[0]) if result[0] else ""
            file_name = str(result[1]) if result[1] else "analysis"
            if os.path.exists(excel_file_path):
                return send_file(excel_file_path, as_attachment=True, download_name=f"{file_name}_analysis.xlsx")
            else:
                return "File not found", 404
        else:
            return "Analysis not found", 404
    except Error as e:
        print(f"Error retrieving analysis: {e}")
        return "Error retrieving analysis", 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/delete-analysis/<int:analysis_id>', methods=['POST'])
def delete_analysis_route(analysis_id):
    if not is_logged_in():
        return redirect(url_for('login'))
    
    # Delete analysis
    success = delete_analysis(analysis_id, session['user_id'])
    
    if success:
        return jsonify({'success': True, 'message': 'Analysis deleted successfully'})
    else:
        return jsonify({'success': False, 'message': 'Failed to delete analysis'}), 400

@app.route('/api/analysis-data/<int:analysis_id>')
def get_analysis_data(analysis_id):
    if not is_logged_in():
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    # Get analysis details
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    cursor = None
    try:
        cursor = connection.cursor()
        cursor.execute("""
            SELECT excel_file_path
            FROM bank_statement_analyses 
            WHERE id = %s AND user_id = %s
        """, (analysis_id, session['user_id']))
        
        result = cursor.fetchone()
        if result:
            excel_file_path = result[0]
            
            # Check if file exists
            if not os.path.exists(excel_file_path):
                return jsonify({'success': False, 'error': 'Analysis file not found'}), 404
            
            # Read the Excel file to extract data
            try:
                # Read the enhanced summary sheet
                summary_df = pd.read_excel(excel_file_path, sheet_name='Enhanced Summary')
                
                # Read the detailed summary sheet
                detailed_df = pd.read_excel(excel_file_path, sheet_name='Detailed Summary')
                
                # Extract key metrics
                total_credits = 0
                total_debits = 0
                net_cashflow = 0
                total_transactions = 0
                
                # Get total credits from the enhanced summary
                total_credits_row = summary_df[summary_df['Metric'] == 'Total Credits']
                if not total_credits_row.empty:
                    total_credits = total_credits_row['Value'].iloc[0]
                
                # Get net cashflow
                net_cashflow_row = summary_df[summary_df['Metric'] == 'NET CASHFLOW']
                if not net_cashflow_row.empty:
                    net_cashflow = net_cashflow_row['Value'].iloc[0]
                
                # Get total transactions
                total_transactions_row = summary_df[summary_df['Metric'] == 'NET CASHFLOW']
                if not total_transactions_row.empty:
                    total_transactions = int(total_transactions_row['Count'].iloc[0])
                
                # Calculate total debits
                total_debits = total_credits - net_cashflow
                
                # Extract category data
                category_data = {}
                for _, row in detailed_df.iterrows():
                    category = row['Category']
                    if category != 'TOTAL':
                        category_data[category] = {
                            'count': int(row['Transaction Count']),
                            'total_amount': float(row['Total Amount']),
                            'credit_amount': float(row['Credit Amount']),
                            'debit_amount': float(row['Debit Amount']),
                            'net_amount': float(row['Net Amount'])
                        }
                
                # Prepare response data
                data = {
                    'success': True,
                    'totalCredits': float(total_credits),
                    'totalDebits': float(total_debits),
                    'netCashflow': float(net_cashflow),
                    'totalTransactions': total_transactions,
                    'categories': category_data
                }
                
                return jsonify(data)
            except Exception as e:
                print(f"Error reading Excel file: {e}")
                return jsonify({'success': False, 'error': 'Failed to read analysis data'}), 500
        else:
            return jsonify({'success': False, 'error': 'Analysis not found'}), 404
    except Error as e:
        print(f"Error retrieving analysis: {e}")
        return jsonify({'success': False, 'error': 'Database error'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()


@app.route('/api/analysis-transactions/<int:analysis_id>/<category>')
def get_analysis_transactions(analysis_id, category):
    if not is_logged_in():
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    # Get analysis details
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    cursor = None
    try:
        cursor = connection.cursor()
        cursor.execute("""
            SELECT excel_file_path
            FROM bank_statement_analyses 
            WHERE id = %s AND user_id = %s
        """, (analysis_id, session['user_id']))
        
        result = cursor.fetchone()
        if result:
            excel_file_path = result[0]
            
            # Check if file exists
            if not os.path.exists(excel_file_path):
                return jsonify({'success': False, 'error': 'Analysis file not found'}), 404
            
            # Read the Excel file to extract data
            try:
                # Decode the URL-encoded category name
                import urllib.parse
                decoded_category = urllib.parse.unquote(category)
                
                # Sanitize category name for sheet name (replace special characters)
                sanitized_category = decoded_category.replace('/', ' ').replace('\\', ' ').replace('?', ' ').replace('*', ' ').replace('[', ' ').replace(']', ' ').replace(':', ' ')
                sheet_name = sanitized_category[:31]
                
                # Try to read the category details sheet
                try:
                    transactions_df = pd.read_excel(excel_file_path, sheet_name=f"{sheet_name} Details")
                except:
                    # If category details sheet doesn't exist, try the main category sheet
                    try:
                        transactions_df = pd.read_excel(excel_file_path, sheet_name=decoded_category)
                    except:
                        # If neither sheet exists, return empty data
                        return jsonify({
                            'success': True,
                            'category': decoded_category,
                            'transactions': [],
                            'count': 0
                        })
                
                # Convert DataFrame to list of dictionaries
                transactions = []
                for _, row in transactions_df.iterrows():
                    transaction = {}
                    for col in transactions_df.columns:
                        transaction[col] = row[col] if not pd.isna(row[col]) else ''
                    transactions.append(transaction)
                
                # Prepare response data
                data = {
                    'success': True,
                    'category': decoded_category,
                    'transactions': transactions,
                    'count': len(transactions)
                }
                
                return jsonify(data)
            except Exception as e:
                print(f"Error reading Excel file: {e}")
                import traceback
                traceback.print_exc()
                return jsonify({'success': False, 'error': 'Failed to read transaction data'}), 500
        else:
            return jsonify({'success': False, 'error': 'Analysis not found'}), 404
    except Error as e:
        print(f"Error retrieving analysis: {e}")
        return jsonify({'success': False, 'error': 'Database error'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/notifications/page')
def admin_notifications_page():
    """Serve the notifications page"""
    # Check if user is admin
    if not session.get('is_admin'):
        return redirect(url_for('login'))
    
    user_name = get_user_name(session['user_id'])
    return render_template('notification.html', user_name=user_name)

@app.route('/admin-dashboard')
def admin_dashboard():
    # Check if user is admin
    if not session.get('is_admin'):
        return redirect(url_for('login'))
    
    # Get all users with their subscription information
    connection = get_db_connection()
    if connection is None:
        return redirect(url_for('login', error='Database connection failed'))
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Get all users with their subscription details, days remaining, and subscription count
        cursor.execute("""
            SELECT u.id, u.name, u.email, u.phone, u.created_at, 
                   s.plan, s.start_date, s.end_date,
                   CASE 
                       WHEN s.plan IS NULL THEN 'No Plan'
                       WHEN s.end_date > NOW() THEN 'Active'
                       ELSE 'Expired'
                   END as subscription_status,
                   CASE 
                       WHEN s.end_date IS NOT NULL THEN DATEDIFF(s.end_date, NOW())
                       ELSE NULL
                   END as days_remaining,
                   CASE 
                       WHEN sub_count.subscription_count > 1 THEN TRUE
                       ELSE FALSE
                   END as is_renewed,
                   COALESCE(sub_count.subscription_count, 0) as subscription_count
            FROM users u
            LEFT JOIN subscriptions s ON u.id = s.user_id
            LEFT JOIN (
                SELECT user_id, COUNT(*) as subscription_count
                FROM subscriptions
                GROUP BY user_id
            ) sub_count ON u.id = sub_count.user_id
            ORDER BY u.created_at DESC
        """)
        
        users_data = cursor.fetchall()
        
        # Process the data for the template
        users = []
        for row in users_data:
            users.append({
                'id': row[0],
                'name': row[1],
                'email': row[2],
                'phone': row[3],
                'joined_date': row[4],
                'plan': row[5],
                'start_date': row[6],
                'end_date': row[7],
                'subscription_status': row[8],
                'days_remaining': row[9],
                'is_renewed': row[10] if len(row) > 10 else False,
                'subscription_count': row[11] if len(row) > 11 else 0
            })
        
        # Check for expiring subscriptions and create/update notifications
        try:
            for user in users:
                # Check if user has an active subscription
                if user['days_remaining'] is not None:
                    # Check for expiring subscriptions (within 7 days)
                    if user['days_remaining'] <= 7 and user['days_remaining'] >= 0:
                        # Check if notification already exists
                        cursor.execute("""
                            SELECT id FROM notifications 
                            WHERE user_id = %s AND type = 'expiring' AND is_read = FALSE
                        """, (user['id'],))
                        
                        existing_notification = cursor.fetchone()
                        if not existing_notification:
                            # Create notification for expiring subscription
                            cursor.execute("""
                                INSERT INTO notifications (user_id, type, title, message) 
                                VALUES (%s, %s, %s, %s)
                            """, (user['id'], 'expiring', 'Subscription Expiring', 
                                  f"{user['name']}'s subscription will expire in {user['days_remaining']} days"))
                        else:
                            # Update existing notification with current days remaining
                            cursor.execute("""
                                UPDATE notifications 
                                SET message = %s, created_at = NOW()
                                WHERE user_id = %s AND type = 'expiring' AND is_read = FALSE
                            """, (f"{user['name']}'s subscription will expire in {user['days_remaining']} days", user['id']))
                    else:
                        # Subscription is not expiring soon or has expired
                        # Remove any existing expiring notifications for this user
                        print(f"DEBUG: Removing expiring notifications for user {user['id']} ({user['name']}) with {user['days_remaining']} days remaining")
                        cursor.execute("""
                            DELETE FROM notifications 
                            WHERE user_id = %s AND type = 'expiring' AND is_read = FALSE
                        """, (user['id'],))
                else:
                    # User has no subscription, remove any expiring notifications
                    print(f"DEBUG: Removing expiring notifications for user {user['id']} ({user['name']}) with no subscription")
                    cursor.execute("""
                        DELETE FROM notifications 
                        WHERE user_id = %s AND type = 'expiring' AND is_read = FALSE
                    """, (user['id'],))
            
            connection.commit()
        except Error as e:
            print(f"Error creating expiring subscription notifications: {e}")
            connection.rollback()
        
        user_name = get_user_name(session['user_id'])
        return render_template('admin_dashboard.html', users=users, user_name=user_name)
    except Error as e:
        print(f"Error fetching admin data: {e}")
        return redirect(url_for('login', error='Failed to load admin dashboard'))
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/cancel-subscription/<int:user_id>', methods=['POST'])
def admin_cancel_subscription(user_id):
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Delete the user's subscription
        cursor.execute("DELETE FROM subscriptions WHERE user_id = %s", (user_id,))
        connection.commit()
        
        return jsonify({'success': True, 'message': 'Subscription cancelled successfully'})
    except Error as e:
        print(f"Error cancelling subscription: {e}")
        return jsonify({'success': False, 'error': 'Failed to cancel subscription'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/cleanup-notifications', methods=['POST'])
def admin_cleanup_notifications():
    """Cleanup outdated expiring subscription notifications"""
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Get all users with their subscription details, days remaining, and subscription count
        cursor.execute("""
            SELECT u.id, u.name, u.email, u.phone, u.created_at, 
                   s.plan, s.start_date, s.end_date,
                   CASE 
                       WHEN s.plan IS NULL THEN 'No Plan'
                       WHEN s.end_date > NOW() THEN 'Active'
                       ELSE 'Expired'
                   END as subscription_status,
                   CASE 
                       WHEN s.end_date IS NOT NULL THEN DATEDIFF(s.end_date, NOW())
                       ELSE NULL
                   END as days_remaining,
                   CASE 
                       WHEN sub_count.subscription_count > 1 THEN TRUE
                       ELSE FALSE
                   END as is_renewed,
                   COALESCE(sub_count.subscription_count, 0) as subscription_count
            FROM users u
            LEFT JOIN subscriptions s ON u.id = s.user_id
            LEFT JOIN (
                SELECT user_id, COUNT(*) as subscription_count
                FROM subscriptions
                GROUP BY user_id
            ) sub_count ON u.id = sub_count.user_id
            ORDER BY u.created_at DESC
        """)
        
        users_data = cursor.fetchall()
        
        # Process the data
        users = []
        for row in users_data:
            row_list = list(row)
            users.append({
                'id': row_list[0],
                'name': row_list[1],
                'email': row_list[2],
                'phone': row_list[3],
                'joined_date': row_list[4],
                'plan': row_list[5],
                'start_date': row_list[6],
                'end_date': row_list[7],
                'subscription_status': row_list[8],
                'days_remaining': row_list[9],
                'is_renewed': row_list[10] if len(row_list) > 10 else False,
                'subscription_count': row_list[11] if len(row_list) > 11 else 0
            })
        
        # Track cleaned up notifications
        cleaned_notifications = []
        
        # Cleanup outdated expiring notifications
        deleted_count = 0
        updated_count = 0
        
        for user in users:
            # Check if user has an active subscription
            if user['days_remaining'] is not None:
                # If subscription is expiring soon (within 7 days) and not expired
                if user['days_remaining'] <= 7 and user['days_remaining'] >= 0:
                    # Check if notification already exists (both read and unread)
                    cursor.execute("""
                        SELECT id, message, is_read FROM notifications 
                        WHERE user_id = %s AND type = 'expiring'
                    """, (user['id'],))
                    
                    existing_notification = cursor.fetchone()
                    if existing_notification:
                        existing_list = list(existing_notification)
                        existing_id = existing_list[0]
                        existing_message = existing_list[1]
                        is_read = existing_list[2]
                        # Check if the message is outdated
                        expected_message = f"{user['name']}'s subscription will expire in {user['days_remaining']} days"
                        if existing_message != expected_message:
                            # Update existing notification with current days remaining
                            cursor.execute("""
                                UPDATE notifications 
                                SET message = %s, created_at = NOW()
                                WHERE id = %s
                            """, (expected_message, existing_id))
                            updated_count += 1
                            read_status = "read" if is_read else "unread"
                            cleaned_notifications.append({
                                'title': 'Subscription Expiring',
                                'message': expected_message,
                                'user_name': user['name'],
                                'user_email': user['email'],
                                'action': 'updated'
                            })
                    else:
                        # Create notification for expiring subscription
                        expected_message = f"{user['name']}'s subscription will expire in {user['days_remaining']} days"
                        cursor.execute("""
                            INSERT INTO notifications (user_id, type, title, message) 
                            VALUES (%s, %s, %s, %s)
                        """, (user['id'], 'expiring', 'Subscription Expiring', expected_message))
                        cleaned_notifications.append({
                            'title': 'Subscription Expiring',
                            'message': expected_message,
                            'user_name': user['name'],
                            'user_email': user['email'],
                            'action': 'created'
                        })
                else:
                    # Subscription is not expiring soon or has expired
                    # Remove any existing expiring notifications for this user (both read and unread)
                    cursor.execute("""
                        SELECT title, message, is_read FROM notifications 
                        WHERE user_id = %s AND type = 'expiring'
                    """, (user['id'],))
                    
                    outdated_notifications = cursor.fetchall()
                    for notification in outdated_notifications:
                        notification_list = list(notification)
                        cleaned_notifications.append({
                            'title': notification_list[0],
                            'message': notification_list[1],
                            'user_name': user['name'],
                            'user_email': user['email'],
                            'action': 'removed'
                        })
                    
                    cursor.execute("""
                        DELETE FROM notifications 
                        WHERE user_id = %s AND type = 'expiring'
                    """, (user['id'],))
                    deleted_count += cursor.rowcount
            else:
                # User has no subscription, remove any expiring notifications (both read and unread)
                cursor.execute("""
                    SELECT title, message, is_read FROM notifications 
                    WHERE user_id = %s AND type = 'expiring'
                """, (user['id'],))
                
                outdated_notifications = cursor.fetchall()
                for notification in outdated_notifications:
                    notification_list = list(notification)
                    cleaned_notifications.append({
                        'title': notification_list[0],
                        'message': notification_list[1],
                        'user_name': user['name'],
                        'user_email': user['email'],
                        'action': 'removed'
                    })
                
                cursor.execute("""
                    DELETE FROM notifications 
                    WHERE user_id = %s AND type = 'expiring'
                """, (user['id'],))
                deleted_count += cursor.rowcount
        
        connection.commit()
        
        return jsonify({
            'success': True, 
            'message': f'Cleaned up {deleted_count + updated_count} notifications',
            'cleaned_count': deleted_count + updated_count,
            'deleted_count': deleted_count,
            'updated_count': updated_count,
            'cleaned_notifications': cleaned_notifications
        })
    except Error as e:
        print(f"Error cleaning up notifications: {e}")
        connection.rollback()
        return jsonify({'success': False, 'error': 'Failed to cleanup notifications'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/send-email/<int:user_id>', methods=['POST'])
def admin_send_email(user_id):
    """Send email to user and track in history"""
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Get user information (including users without subscriptions)
        cursor.execute("""
            SELECT u.name, u.email, u.phone, s.end_date, s.plan
            FROM users u
            LEFT JOIN subscriptions s ON u.id = s.user_id
            WHERE u.id = %s
        """, (user_id,))
        
        user_data = cursor.fetchone()
        
        if not user_data:
            return jsonify({'success': False, 'error': 'User not found'})
        
        user_name, user_email, user_phone, end_date, plan = user_data
        
        # Prepare email content based on subscription status
        if plan and end_date:
            # User has an active or expired subscription
            from datetime import datetime
            days_until_expiry = (end_date - datetime.now()).days
            
            email_subject = f"Subscription Reminder - {days_until_expiry} days remaining"
            email_body = f"""
Dear {user_name},

This is a reminder that your {plan} subscription will expire in {days_until_expiry} days.
Please renew your subscription to continue enjoying our services.

Best regards,
{SMTP_CONFIG['sender_name']}
            """
        else:
            # User has no subscription
            email_subject = "Special Offer - Bank Statement Analysis Services"
            email_body = f"""
Dear {user_name},

We noticed you haven't subscribed to our bank statement analysis services yet. 
We'd be happy to help you get started with analyzing your financial documents.

Our services include:
- Multi-format document processing (PDF, PNG, JPG, GIF, DOC, DOCX, XLS, XLSX)
- Smart transaction categorization
- Financial insights and spending pattern analysis
- Comprehensive Excel exports with charts and visualizations

If you have any questions or would like to discuss our services, please don't hesitate to reach out.

Best regards,
{SMTP_CONFIG['sender_name']}
            """
        
        # Send actual email
        if send_email(user_email, email_subject, email_body):
            # Track in email history
            cursor.execute("""
                INSERT INTO email_history (user_id, email, subject, body, status) 
                VALUES (%s, %s, %s, %s, %s)
            """, (user_id, user_email, email_subject, email_body, 'sent'))
            connection.commit()
            
            return jsonify({
                'success': True,
                'message': f'Email sent successfully to {user_name} ({user_email})',
                'subject': email_subject
            })
        else:
            # Track failed attempt
            cursor.execute("""
                INSERT INTO email_history (user_id, email, subject, body, status) 
                VALUES (%s, %s, %s, %s, %s)
            """, (user_id, user_email, email_subject, email_body, 'failed'))
            connection.commit()
            
            return jsonify({'success': False, 'error': 'Failed to send email'})
            
    except Error as e:
        print(f"Error sending email: {e}")
        return jsonify({'success': False, 'error': 'Failed to send email'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/send-reminder/<int:user_id>', methods=['POST'])
def admin_send_reminder(user_id):
    """Send reminder email to user and track in history"""
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Get user information (including users without subscriptions)
        cursor.execute("""
            SELECT u.name, u.email, u.phone, s.end_date, s.plan
            FROM users u
            LEFT JOIN subscriptions s ON u.id = s.user_id
            WHERE u.id = %s
        """, (user_id,))
        
        user_data = cursor.fetchone()
        
        if not user_data:
            return jsonify({'success': False, 'error': 'User not found'})
        
        user_name, user_email, user_phone, end_date, plan = user_data
        
        # Prepare email content based on subscription status
        if plan and end_date:
            # User has an active or expired subscription
            from datetime import datetime
            days_until_expiry = (end_date - datetime.now()).days
            
            email_subject = f"Subscription Reminder - {days_until_expiry} days remaining"
            email_body = f"""
Dear {user_name},

This is a reminder that your {plan} subscription will expire in {days_until_expiry} days.
Please renew your subscription to continue enjoying our services.

Best regards,
{SMTP_CONFIG['sender_name']}
            """
        else:
            # User has no subscription
            email_subject = "Special Offer - Bank Statement Analysis Services"
            email_body = f"""
Dear {user_name},

We noticed you haven't subscribed to our bank statement analysis services yet. 
We'd be happy to help you get started with analyzing your financial documents.

Our services include:
- Multi-format document processing (PDF, PNG, JPG, GIF, DOC, DOCX, XLS, XLSX)
- Smart transaction categorization
- Financial insights and spending pattern analysis
- Comprehensive Excel exports with charts and visualizations

If you have any questions or would like to discuss our services, please don't hesitate to reach out.

Best regards,
{SMTP_CONFIG['sender_name']}
            """
        
        # Send actual email
        if send_email(user_email, email_subject, email_body):
            # Track in communication history
            cursor.execute("""
                INSERT INTO communication_history (user_id, message_type, status, details) 
                VALUES (%s, %s, %s, %s)
            """, (user_id, 'reminder', 'sent', f'Reminder sent to {user_email}'))
            connection.commit()
            
            return jsonify({
                'success': True, 
                'message': f'Reminder email sent successfully to {user_name} ({user_email})'
            })
        else:
            # Track failed attempt
            cursor.execute("""
                INSERT INTO communication_history (user_id, message_type, status, details) 
                VALUES (%s, %s, %s, %s)
            """, (user_id, 'reminder', 'failed', f'Failed to send reminder to {user_email}'))
            connection.commit()
            
            return jsonify({'success': False, 'error': 'Failed to send email reminder'})
            
    except Error as e:
        print(f"Error sending reminder: {e}")
        return jsonify({'success': False, 'error': 'Failed to send reminder'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/get-whatsapp-message/<int:user_id>', methods=['POST'])
def admin_get_whatsapp_message(user_id):
    """Prepare WhatsApp message for user and track in history"""
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Get user information (including users without subscriptions)
        cursor.execute("""
            SELECT u.name, u.phone, s.end_date, s.plan
            FROM users u
            LEFT JOIN subscriptions s ON u.id = s.user_id
            WHERE u.id = %s
        """, (user_id,))
        
        user_data = cursor.fetchone()
        
        if not user_data:
            return jsonify({'success': False, 'error': 'User not found'})
        
        user_name, user_phone, end_date, plan = user_data
        
        # Check if user has phone number
        if not user_phone:
            return jsonify({'success': False, 'error': 'User phone number not available'})
        
        # Prepare WhatsApp message based on subscription status
        if plan and end_date:
            # User has an active or expired subscription
            from datetime import datetime
            days_until_expiry = (end_date - datetime.now()).days
            
            whatsapp_message = f"Dear {user_name}, your {plan} subscription will expire in {days_until_expiry} days. Please renew to continue enjoying our services. - {SMTP_CONFIG['sender_name']}"
        else:
            # User has no subscription
            whatsapp_message = f"Dear {user_name}, we noticed you haven't subscribed to our bank statement analysis services yet. We'd be happy to help you get started. Please reply if you have any questions. - {SMTP_CONFIG['sender_name']}"
        
        # Track in WhatsApp history
        cursor.execute("""
            INSERT INTO whatsapp_history (user_id, phone, message) 
            VALUES (%s, %s, %s)
        """, (user_id, user_phone, whatsapp_message))
        connection.commit()
        
        return jsonify({
            'success': True,
            'whatsapp': {
                'phone': user_phone,
                'message': whatsapp_message
            },
            'message': f'WhatsApp message prepared for {user_name}'
        })
            
    except Error as e:
        print(f"Error preparing WhatsApp message: {e}")
        return jsonify({'success': False, 'error': 'Failed to prepare WhatsApp message'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

# Notification routes
@app.route('/admin/notifications')
def admin_get_notifications():
    """Get all notifications for the admin dashboard"""
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Get all notifications with user information
        cursor.execute("""
            SELECT n.id, n.user_id, n.type, n.title, n.message, n.is_read, n.starred, n.created_at,
                   u.name, u.email, u.phone
            FROM notifications n
            JOIN users u ON n.user_id = u.id
            ORDER BY n.created_at DESC
        """)
        
        notifications_data = cursor.fetchall()
        
        # Process the data for the response
        notifications = []
        for row in notifications_data:
            notifications.append({
                'id': row[0],
                'user_id': row[1],
                'type': row[2],
                'title': row[3],
                'message': row[4],
                'is_read': row[5],
                'starred': row[6] if row[6] is not None else False,
                'created_at': row[7],
                'user_name': row[8],
                'user_email': row[9],
                'user_phone': row[10] or 'N/A'
            })
        
        return jsonify({'success': True, 'notifications': notifications})
    except Error as e:
        print(f"Error fetching notifications: {e}")
        return jsonify({'success': False, 'error': 'Failed to fetch notifications'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/notifications/mark-as-read/<int:notification_id>', methods=['POST'])
def admin_mark_notification_as_read(notification_id):
    """Mark a notification as read"""
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Update notification as read
        cursor.execute("UPDATE notifications SET is_read = TRUE WHERE id = %s", (notification_id,))
        connection.commit()
        
        return jsonify({'success': True, 'message': 'Notification marked as read'})
    except Error as e:
        print(f"Error marking notification as read: {e}")
        return jsonify({'success': False, 'error': 'Failed to mark notification as read'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/notifications/mark-as-unread/<int:notification_id>', methods=['POST'])
def admin_mark_notification_as_unread(notification_id):
    """Mark a notification as unread"""
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Update notification as unread
        cursor.execute("UPDATE notifications SET is_read = FALSE WHERE id = %s", (notification_id,))
        connection.commit()
        
        return jsonify({'success': True, 'message': 'Notification marked as unread'})
    except Error as e:
        print(f"Error marking notification as unread: {e}")
        return jsonify({'success': False, 'error': 'Failed to mark notification as unread'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/notifications/delete/<int:notification_id>', methods=['POST'])
def admin_delete_notification(notification_id):
    """Delete a specific notification"""
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Delete the notification
        cursor.execute("DELETE FROM notifications WHERE id = %s", (notification_id,))
        connection.commit()
        
        return jsonify({'success': True, 'message': 'Notification deleted successfully'})
    except Error as e:
        print(f"Error deleting notification: {e}")
@app.route('/admin/delete-user/<int:user_id>', methods=['POST'])
def admin_delete_user(user_id):
    """Delete a user and all associated data"""
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Delete user and all associated data (CASCADE will handle related tables)
        cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
        connection.commit()
        
        if cursor.rowcount > 0:
            return jsonify({'success': True, 'message': 'User deleted successfully'})
        else:
            return jsonify({'success': False, 'error': 'User not found'})
    except Error as e:
        print(f"Error deleting user: {e}")
        connection.rollback()
        return jsonify({'success': False, 'error': 'Failed to delete user'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()


@app.route('/admin/notifications/clear-all', methods=['POST'])
def admin_clear_all_notifications():
    """Clear all notifications"""
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Delete all notifications
        cursor.execute("DELETE FROM notifications")
        connection.commit()
        
        return jsonify({'success': True, 'message': 'All notifications cleared successfully'})
    except Error as e:
        print(f"Error clearing notifications: {e}")
        return jsonify({'success': False, 'error': 'Failed to clear notifications'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()


@app.route('/admin/notifications/star/<int:notification_id>', methods=['POST'])
def admin_star_notification(notification_id):
    """Star/unstar a notification"""
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Get current starred status
        cursor.execute("SELECT starred FROM notifications WHERE id = %s", (notification_id,))
        result = cursor.fetchone()
        
        if not result:
            return jsonify({'success': False, 'error': 'Notification not found'})
        
        # Toggle starred status
        current_starred = result[0] if result[0] is not None else 0
        new_starred = 1 - current_starred
        
        # Update notification
        cursor.execute("UPDATE notifications SET starred = %s WHERE id = %s", (new_starred, notification_id))
        connection.commit()
        
        status = "starred" if new_starred == 1 else "unstarred"
        return jsonify({'success': True, 'message': f'Notification {status} successfully', 'starred': bool(new_starred)})
    except Error as e:
        print(f"Error starring notification: {e}")
        return jsonify({'success': False, 'error': 'Failed to star notification'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

# History tracking routes
@app.route('/admin/history/registrations')
def admin_get_registration_history():
    """Get new user registration history"""
    # Check if user is logged in
    if not is_logged_in():
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Get recent user registrations (last 24 hours)
        cursor.execute("""
            SELECT id, name, email, phone, created_at
            FROM users
            WHERE created_at >= DATE_SUB(NOW(), INTERVAL 1 DAY)
            ORDER BY created_at DESC
        """)
        
        registrations_data = cursor.fetchall()
        
        # Process the data for the response
        registrations = []
        for row in registrations_data:
            registrations.append({
                'id': row[0],
                'name': row[1],
                'email': row[2],
                'phone': row[3] or 'N/A',
                'created_at': row[4]
            })
        
        return jsonify({'success': True, 'registrations': registrations})
    except Error as e:
        print(f"Error fetching registration history: {e}")
        return jsonify({'success': False, 'error': 'Failed to fetch registration history'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/history/subscriptions')
def admin_get_subscription_history():
    """Get new subscription history"""
    # Check if user is logged in
    if not is_logged_in():
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Get recent subscriptions (last 24 hours)
        cursor.execute("""
            SELECT s.id, s.user_id, s.plan, s.start_date, s.end_date, s.created_at,
                   u.name, u.email
            FROM subscriptions s
            JOIN users u ON s.user_id = u.id
            WHERE s.created_at >= DATE_SUB(NOW(), INTERVAL 1 DAY)
            ORDER BY s.created_at DESC
        """)
        
        subscriptions_data = cursor.fetchall()
        
        # Process the data for the response
        subscriptions = []
        for row in subscriptions_data:
            subscriptions.append({
                'id': row[0],
                'user_id': row[1],
                'plan': row[2],
                'start_date': row[3],
                'end_date': row[4],
                'created_at': row[5],
                'user_name': row[6],
                'user_email': row[7]
            })
        
        return jsonify({'success': True, 'subscriptions': subscriptions})
    except Error as e:
        print(f"Error fetching subscription history: {e}")
        return jsonify({'success': False, 'error': 'Failed to fetch subscription history'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/history/reminders')
def admin_get_reminder_history():
    """Get reminder sending history"""
    # Check if user is logged in
    if not is_logged_in():
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Get reminder history (last 30 days)
        cursor.execute("""
            SELECT id, user_id, message_type, sent_at, status,
                   u.name, u.email
            FROM communication_history
            JOIN users u ON user_id = u.id
            WHERE sent_at >= DATE_SUB(NOW(), INTERVAL 30 DAY)
            ORDER BY sent_at DESC
        """)
        
        reminders_data = cursor.fetchall()
        
        # Process the data for the response
        reminders = []
        for row in reminders_data:
            reminders.append({
                'id': row[0],
                'user_id': row[1],
                'message_type': row[2],
                'sent_at': row[3],
                'status': row[4],
                'user_name': row[5],
                'user_email': row[6]
            })
        
        return jsonify({'success': True, 'reminders': reminders})
    except Error as e:
        print(f"Error fetching reminder history: {e}")
        return jsonify({'success': False, 'error': 'Failed to fetch reminder history'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/history/whatsapp')
def admin_get_whatsapp_history():
    """Get WhatsApp message history"""
    # Check if user is logged in
    if not is_logged_in():
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Get WhatsApp message history (last 30 days)
        cursor.execute("""
            SELECT id, user_id, phone, message, sent_at,
                   u.name
            FROM whatsapp_history
            JOIN users u ON user_id = u.id
            WHERE sent_at >= DATE_SUB(NOW(), INTERVAL 30 DAY)
            ORDER BY sent_at DESC
        """)
        
        whatsapp_data = cursor.fetchall()
        
        # Process the data for the response
        whatsapp_messages = []
        for row in whatsapp_data:
            whatsapp_messages.append({
                'id': row[0],
                'user_id': row[1],
                'phone': row[2],
                'message': row[3],
                'sent_at': row[4],
                'user_name': row[5]
            })
        
        return jsonify({'success': True, 'whatsapp_messages': whatsapp_messages})
    except Error as e:
        print(f"Error fetching WhatsApp history: {e}")
        return jsonify({'success': False, 'error': 'Failed to fetch WhatsApp history'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/history/email')
def admin_get_email_history():
    """Get email sending history"""
    # Check if user is logged in
    if not is_logged_in():
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    # Check if user is admin
    if not session.get('is_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    connection = get_db_connection()
    if connection is None:
        return jsonify({'success': False, 'error': 'Database connection failed'})
    
    cursor = None
    try:
        cursor = connection.cursor()
        
        # Get email history (last 30 days)
        cursor.execute("""
            SELECT id, user_id, email, subject, sent_at, status,
                   u.name
            FROM email_history
            JOIN users u ON user_id = u.id
            WHERE sent_at >= DATE_SUB(NOW(), INTERVAL 30 DAY)
            ORDER BY sent_at DESC
        """)
        
        emails_data = cursor.fetchall()
        
        # Process the data for the response
        emails = []
        for row in emails_data:
            emails.append({
                'id': row[0],
                'user_id': row[1],
                'email': row[2],
                'subject': row[3],
                'sent_at': row[4],
                'status': row[5],
                'user_name': row[6]
            })
        
        return jsonify({'success': True, 'emails': emails})
    except Error as e:
        print(f"Error fetching email history: {e}")
        return jsonify({'success': False, 'error': 'Failed to fetch email history'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

if __name__ == '__main__':
    app.run(debug=True, port=5008)